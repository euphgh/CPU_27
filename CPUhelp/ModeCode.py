import openpyxl
def alwaysgenerator(port_head,port_bottom,sheet,modulename):
    forward = ""
    initial = ""
    defines = ""
    assign = ""
    reg = ""
    wire = ""
    for j in range(port_head, port_bottom):
        if sheet.cell(j,6).value == "datar" or sheet.cell(j,6).value=="dataw":
            varname = sheet.cell(j,2).value[:-3]
            x = varname.find("_")
            prefix = modulename.lower() + "_"
            innername = varname[:x+1] + "to_" +prefix + varname[x+1:]
            if sheet.cell(j,6).value == "datar":
                forward = forward + "\n\t" + "{0}_r <= {1}_in;".format(innername,varname)
                initial = initial + "\n\t" + "{0}_r <= `ini_{1}_in;".format(innername,varname)
                defines = defines + "\n" +"`define ini_{0}_in {1}'b0".format(varname,sheet.cell(j,3).value)
                if sheet.cell(j,3).value > 1:
                    reg = reg + "reg  [{0}:0] {1}_r ;\n".format(sheet.cell(j,3).value-1,innername)
                else:
                    reg = reg + "reg  {0}_r ;\n".format(innername)
            else:
                assign = assign + "assign {0}_r = {1}_in ;\n".format(innername,varname)
                if sheet.cell(j,3).value > 1:
                    wire = wire + "wire [{0}:0] {1}_r ;\n".format(sheet.cell(j,3).value-1,innername)
                else:
                    wire = wire + "wire {0}_r ;\n".format(innername)

    always = '''\
always @(posedge clk) begin
    if (!rst_n) begin{0}
    end 
    else if (allowin && valid_r) begin{1}
    end
end'''.format(initial,forward)
    print(defines)
    print(wire)
    print(reg)
    print(assign)
    print(always)

def AlwaysCode(updatelist,book):
    for i in updatelist:
        sheet = book[i]
        name = i
        port_head = 3
        port_bottom = port_head
        for j in range(1, sheet.max_row+1):
            if sheet.cell(j, 1).value == "variable":
                port_bottom = j
                break
        alwaysgenerator(port_head,port_bottom,sheet,i)
        
book = openpyxl.load_workbook("CPU-26.xlsx")
gl = ["WB"]
AlwaysCode(gl,book)
book.close()