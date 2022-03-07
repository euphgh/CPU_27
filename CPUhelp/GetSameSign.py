import openpyxl as pxl
from numpy import *
book = pxl.load_workbook("CPU-26.xlsx")
sheet = book["sign"]
colmax = pxl.utils.column_index_from_string("BE")
def contrast(signlist,conlist):
    for keyrow in conlist:
        samelist = []
        less = 4
        lessrow = keyrow
        lesskind = ""
        for row in signlist:
            if row==keyrow:continue
            samelist = [sheet.cell(row,col).value == sheet.cell(keyrow,col).value for col in range(2,colmax+1)]
            samenum = samelist.count(True)
            diffnum = (colmax-1) - samenum
            if samenum<=3:
                if samenum<less:
                    less = samenum
                    lessrow = row
                    lesskind = "diff"
            elif diffnum<=3:
                if diffnum<less:
                    less = diffnum
                    lessrow = row
                    lesskind = "same"
        if lesskind=="same":
            print("{0} is similar with {1} , but has {2} difference.".format(sheet.cell(keyrow,1).value,sheet.cell(lessrow,1).value,less))
        else :
            print("{0} is diff with {1} , but has {2} same.".format(sheet.cell(keyrow,1).value,sheet.cell(lessrow,1).value,less))

signlist = arange(35,101)
conlist = arange(35,101)
contrast(signlist,conlist)