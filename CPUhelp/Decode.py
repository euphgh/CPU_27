from asyncore import read
from typing import Tuple
import openpyxl as pxl
import copy
from numpy import *

book = pxl.load_workbook("CPU-26.xlsx")
sheet = book["sign"]
codefile=open("signcode.txt","wt",encoding="utf-8")
infofile=open("info.txt","wt",encoding="utf-8") 
signRow = lambda x: x+35
lutnum = 0

def binary(bit):

    res = []
    for i in range(2**bit):
        temp = list(bin(i))[2:]
        num = len(temp)
        for j in range(bit):
            if j<=num-1:
                temp[j] = int(temp[j])
            else:
                temp.insert(0,0)
        res.append(temp)
    res.sort(key=lambda x:x.count(1))
    res = array(res[1:],dtype=int)
    return res

def dictest(idict,group):
    pos = where(group==2)[0]
    if pos.size == 0:
        # print("idictest:",group)
        if idict.get(tuple(group),0)==0:return True
        else: return False
    else:
        temp0 = group.copy()
        temp0[pos[0]] = 0
        if  dictest(idict,temp0)==False:
            return False
        else:
            temp1 = group.copy()
            temp1[pos[0]]=1
            if dictest(idict,temp1)==False:
                return False
        return True

def addcodeline(idict,ena,enlist,name):
    global lutnum
    lutnum += ((int)((sum(ena))-1))//6+1
    codeline = "assign {} = ".format(name)
    oritem = ""
    n = 0
    for period in idict.keys():#遍历字典中每一个元素
        if idict[period]==0:continue
        anditem = ""    
        for biten in range(len(ena)):
            if period[biten]==1:
                anditem =anditem + "&&" + "ins[{}]".format(enlist[biten])
            elif ena[biten]==1:
                anditem =anditem + "&&" + "(!ins[{}])".format(enlist[biten])
        anditem = anditem[2:]
        oritem = oritem+" || "+anditem##生成或项
        n = n + 1
        if n%3==0:
            oritem = oritem+"\n"
    oritem = oritem[4:]

    if oritem[-1]=="\n":
        codeline = codeline + oritem[:-1] +" ;\n"
    else:
        codeline = codeline + oritem + " ;\n"
    codefile.write(codeline)
    return n

def DictAddTest(idict,group,val):
    pos = where(group==2)[0]
    tgroup = tuple(group)
    if pos.size == 0:##递归终止
        if idict.get(tgroup,2)==2:
            idict[tgroup] = val
            return idict,True
        else:
            if idict[tgroup]==val:
                return idict,True 
            else:
                return idict,False
    else:##递归
        temp0 = group.copy()
        temp0[pos[0]] = 0
        nidict,flag0 = DictAddTest(idict,temp0,val)
        if flag0 == False:
            return nidict,False
        else:
            temp1 = group.copy()
            temp1[pos[0]]=1
            nidict,flag1 = DictAddTest(idict,temp1,val)
            if flag1 == False:
                return nidict,False
        return nidict,True

def writecode(level,row_sign,idict,ena,bitable):
    fullname =  (sheet.cell(row_sign,1).value)
    gap = fullname.find("/")
    name = fullname[:gap]
    signdict[row_sign]=1;
    num =  addcodeline(idict,ena,bitable,name)
    infofile.write("Sucess information ({0}):\nname:{1}\nenarow:{2}\nactena:{3}\nAndItemNum:{4}\n".format(level,fullname,bitable,ena,num))

def LowLevel(bitable,colmax,signdict):
    torow = lambda x:34-x
    bitable = (31,30,29,28,27,26)
    binarylist = binary(len(bitable))
    for row_sign in signdict.keys():
        if signdict[row_sign]==1: continue
        for ena in binarylist:##存放那几行有效
            idict = {}
            flag = True
            for col in range(2,colmax+1):
                temp = []
                bitval = sheet.cell(row_sign, col).value
                temp = [sheet.cell(torow(i),col).value for i in bitable]
                temp = array(temp*ena,dtype=int)
                # print(temp)
                idict,flag = DictAddTest(idict,temp,bitval)
                # print(idict,flag)
                if flag==False:break
            if flag==True:
                writecode("Low",row_sign,idict,ena,bitable)
                break
            if sum(ena)==len(ena):
                fullname =  (sheet.cell(row_sign,1).value)
                infofile.write("Fail Information (Low):\nname:{}\n".format(fullname))


def MiddleLevel(bitable,colmax,signdict):
    torow = lambda x:34-x
    line = colmax
    for row_sign in signdict.keys():
        if signdict[row_sign]==1: continue
        enlist = list(bitable)
        effectcol = []##存放哪几列有效
        for col_instr in range(2,line+1):
            bitval = sheet.cell(row_sign,col_instr).value
            if bitval==1:
                effectcol.append(col_instr)
                removelist = []
                for num in enlist:
                    if sheet.cell(torow(num),col_instr).value!=0 and sheet.cell(torow(num),col_instr).value!=1:
                        removelist.append(num)
                for num in removelist:
                    enlist.remove(num)
                    
        enlist = array(enlist,dtype=int)
        binarylist = binary(len(enlist))
        for ena in binarylist:##存放那几行有效
            instruct = {}
            for effcol in effectcol:##将1的列加入字典
                temp = array([sheet.cell(torow(i),effcol).value for i in enlist],dtype=int)
                temp = tuple(temp * ena)
                instruct[temp] = 1
            # print(ena,instruct)

            flag = True
            col = 2
            while flag and col<=colmax:
                bitval = sheet.cell(row_sign, col).value
                if bitval==1:
                    col += 1
                    continue
                temp=[]
                for i in enlist: ##生成0列，含有“2”的列表
                    tempval = sheet.cell(torow(i),col).value
                    if tempval==0 or tempval==1:
                        temp.append(tempval)
                    else:
                        temp.append(2)
                temp = array(temp*ena,dtype=int)
                if dictest(instruct,temp)==False:
                    flag = False
                    break
                col += 1

            if flag:
                writecode("Middle",row_sign,instruct,ena,enlist)
                break
            if sum(ena)==len(ena):
                fullname =  (sheet.cell(row_sign,1).value)
                infofile.write("Fail Information (Middel):\nname:{}\n".format(fullname))


def HighLevel(bitable,colmax,signdict):
    torow = lambda x:34-x
    binarylist = binary(len(bitable))
    for row_sign in signdict.keys():
        if signdict[row_sign]==1: continue
        for ena in binarylist:##存放那几行有效
            idict = {}
            flag = True
            for col in range(2,colmax+1):
                temp = []
                bitval = sheet.cell(row_sign, col).value
                for i in bitable: ##生成列，含有“2”的列表
                    tempval = sheet.cell(torow(i),col).value
                    if tempval==0 or tempval==1:
                        temp.append(tempval)
                    else:
                        temp.append(2)
                temp = array(temp*ena,dtype=int)
                # print(temp)
                idict,flag = DictAddTest(idict,temp,bitval)
                # print(idict,flag)
                if flag==False:break
            if flag==True:
                writecode("Hign",row_sign,idict,ena,bitable)
                break
            if sum(ena)==len(ena):
                fullname =  (sheet.cell(row_sign,1).value)
                infofile.write("Fail Information (High):\nname:{}\n".format(fullname))

def check(signdict):
    flag = True
    for i in signdict.items():
        if i[1]==0:
            print(i[0],end=" ")
            flag = False
    if flag: 
        print("All signs expressions generate successfully!")
    else:
        print("Above signs expressions fail to  generate!")
        print(signdict)

# def updateDecoder():
#     vfile = open("../decoder.v","rt",encoding="utf-8")
#     while True:
#         line = vfile.readline()
#         if line=="":
#             break
#         if line.strip()=="//--------------------Python Code--------------------":
            

bitable = (31,30,29,28,27,26,20,16,5,4,3,2,1,0)
signlist = [88,89] 
signdict = dict([(i,0) for i in signlist])
print(signdict)
colmax = pxl.utils.column_index_from_string("BE")
LowLevel(bitable,colmax,signdict)
MiddleLevel(bitable,colmax,signdict)
HighLevel(bitable,colmax,signdict)
check(signdict)
book.close()
codefile.close()
infofile.close()
print("LUT number is:",lutnum)

