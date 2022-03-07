import openpyxl as pxl

book = pxl.load_workbook("CPU-26.xlsx")
def getname(temp):
    name = []
    pre = 0
    for j in range(len(temp)):
        if temp[j] == "," or temp[j] == ";":
            name.append(temp[pre:j])
            pre = j + 1
    if pre == 0:
        name.append(temp)
    return name


def getother(other):
    pos = other.find(":")
    usage = other[:pos]
    explain = other[pos + 1:]
    return usage, explain


def cleanlast(lastline,moduleName):
    sheet = book[moduleName]
    if sheet.max_row > lastline:
        for i in range(lastline, sheet.max_row + 1):
            for j in range(1, 10):
                sheet.cell(i, j).value = None
    book.save("CPU-26.xlsx")


def writeback(data, moduleName, startIndex, typeName):
    namelist = book.sheetnames
    flag = False
    for i in namelist:
        if i == moduleName:
            flag = True
            sheet = book[moduleName]
            break
    if not (flag):
        sheet = book.create_sheet(moduleName)
    sheet.cell(startIndex, 1).value = typeName
    for i in range(2, 6):
        sheet.cell(startIndex, i).value = None
    startIndex += 1
    for line in data:
        for j in range(len(line)):
            sheet.cell(startIndex, j + 1).value = line[j]
        startIndex += 1
    book.save("CPU-26.xlsx")
    return startIndex


def updateExcel(moduleName):
    file = open("../rtl/{}.v".format(moduleName), "rt", encoding="utf-8")
    flag = True
    n = 0
    linelist = []
    portStartLine = 0
    portEndLine = portStartLine
    while flag:
        temp = file.readline()
        if temp=="\n":
            continue
        else:
            linelist.append(temp.strip().split())
        if linelist[n][0] == "/*====================Ports":
            portStartLine = n
        elif linelist[n][0] == "/*====================Variable":
            portEndLine = n
            flag = False
        elif linelist[n] == "":
            flag = False
        n += 1

    df = [["number", "name", "bit", "in/out", "type",  "usage", "expalin"]]
    number = 1
    for i in range(portStartLine, portEndLine):
        if linelist[i][0] == "input" or linelist[i][0] == "output":
            temp = linelist[i][2]
            usage = None
            explain = None
            if temp[0] == "[":
                bit = int(temp[1:temp.find(":")]) + 1
                name = getname(linelist[i][3])
                if len(linelist[i]) > 3:
                    other = "".join(linelist[i][4:])[2:]
                usage, explain = getother(other)
            else:
                bit = 1
                name = getname(temp)
                if len(linelist[i]) > 2:
                    other = ("".join(linelist[i][3:]))[2:]
                usage, explain = getother(other)
            for k in range(len(name)):
                df.append([number, name[k], bit, linelist[i][0][:-3], linelist[i][1], usage, explain])
                number += 1
    for i in df:
        print(i)
    lastIndex = writeback(df, moduleName, 1, "port")
    linelist = []
    n = 0
    flag = True
    varStartLine = 0
    varEndLine = varStartLine
    while flag:
        temp = file.readline()
        if temp=="\n":
            continue
        else:
            linelist.append(temp.strip().split())
        if linelist[n][0] == "/*====================Variable":
            varStartLine = n
        elif linelist[n][0] == "/*====================Function":
            varEndLine = n
            flag = False
        elif linelist[n] == "":
            flag = False
        n += 1
    print(varStartLine,varEndLine)
    df = [["number", "name", "bit", "in/out", "type",  "usage", "expalin"]]
    number = 1
    for i in range(varStartLine, varEndLine):
        if linelist[i][0] == "reg" or linelist[i][0] == "wire":
            temp = linelist[i][1]
            usage = None
            explain = None
            if temp[0] == "[":
                bit = int(temp[1:temp.find(":")]) + 1
                name = getname(linelist[i][2])
                if len(linelist[i]) > 2:
                    other = ("".join(linelist[i][3:]))[2:]
                usage, explain = getother(other)
            else:
                bit = 1
                name = getname(temp)
                if len(linelist[i]) > 1:
                    other = ("".join(linelist[i][2:]))[2:]
                usage, explain = getother(other)
            for k in range(len(name)):
                df.append([number, name[k], bit, None, linelist[i][0], usage, explain])
                number += 1
    for i in df:
        print(i)            
    lastIndex = writeback(df, moduleName, lastIndex, "variable")
    cleanlast(lastIndex,moduleName)
    file.close()

moduleName = ["WB"]
for i in moduleName:
    updateExcel(i)
book.close()
print("Successfully update the excel.")