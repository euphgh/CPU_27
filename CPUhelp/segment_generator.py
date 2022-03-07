import openpyxl as pxl
import re
def prefixgenerator(port_head,port_bottom,name,sheet):
    prefix = "module {}(\n \
\t);".format(name)
    for j in range(port_head, port_bottom):
        line = ""
        line = line + ("input " if sheet.cell(j, 4).value == "in" else "output ")
        line = line + ("wire " if sheet.cell(j, 5).value == "wire" else "reg ")
        if sheet.cell(j, 3).value != None:
            if sheet.cell(j, 3).value != "1" or sheet.cell(j, 3).value != 1:
                line = line + "[{}:0] ".format(int(sheet.cell(j, 3).value) - 1)
        line = line + sheet.cell(j, 2).value + ","
        if j == port_bottom - 1:
            line = line[:-1]
        line = line + " //" + ("Warning:no bit in excel! " if sheet.cell(j, 3).value == None else "") \
               + ((sheet.cell(j, 6).value) if sheet.cell(j, 6).value != None else "") \
               + ((":{}".format(sheet.cell(j, 7).value)) if sheet.cell(j, 7).value != None else "")
        prefix = prefix[:-3] + "\t{}\n".format(line) + prefix[-3:]
    return prefix

def Variablegenerator(inside_head,inside_bottom,sheet):
    Variable = ""
    for j in range(inside_head,inside_bottom):
        line = ""
        line = line + ("wire " if sheet.cell(j, 5).value == "wire" else "reg ")
        if sheet.cell(j, 3).value != None or sheet.cell(j, 3).value != "":
            print(sheet.cell(j, 3).value)
            if sheet.cell(j, 3).value != "1" or sheet.cell(j, 3).value != 1:
                line = line + "[{}:0] ".format(int(sheet.cell(j, 3).value) - 1)
        line = line + sheet.cell(j, 2).value + ";"
        line = line + " //" + ("Warning:no bit in excel! " if sheet.cell(j, 3).value == None else "") \
               + ((sheet.cell(j, 6).value) if sheet.cell(j, 6).value != None else "") \
               + (":{}".format(sheet.cell(j, 7).value) if (sheet.cell(j, 7).value != "") else "")
        Variable = Variable + line + "\n"
    return Variable

def GenerateCode(generatelist,book):
    for i in generatelist:
        sheet = book[i]
        name = i
        port_head = 3
        port_bottom = port_head
        inside_bottom = sheet.max_row
        inside_head = inside_bottom
        for j in range(1, inside_bottom):
            if sheet.cell(j, 1).value == "variable":
                port_bottom = j
                inside_head = j + 2
                break
        print(port_head,port_bottom,inside_head,inside_bottom)
        prefix = prefixgenerator(port_head, port_bottom,name,sheet)
        code = "/*====================Ports Declaration====================*/\n" + prefix + "\n"
        Variable = Variablegenerator(inside_head, inside_bottom,sheet)
        code = code + \
               "\n/*====================Variable Declaration====================*/\n" + Variable +"\n"
        code = code + \
               "/*====================Function Code====================*/\n"
        code = code+"\nendmodule"
        file =  open("../{}.v".format(name),mode="wt+",encoding="utf-8")
        data = file.read()
        pos = re.search("/\*====================Function Code====================\*/", data)
        if pos==None:
            pos = 0
        else:
            pos = pos.span[0]
        data = code + data[pos:]
        file.close()
        file =  open("../{}.v".format(name),mode="wt+",encoding="utf-8")
        file.write(data)
        print(data)
        file.close()

def UpdateCode(updatelist,book):
    for i in updatelist:
        sheet = book[i]
        name = i
        port_head = 3
        port_bottom = port_head
        inside_bottom = sheet.max_row + 1
        inside_head = inside_bottom
        for j in range(1, inside_bottom):
            if sheet.cell(j, 1).value == "variable":
                port_bottom = j
                inside_head = j + 2
                break
        print(port_head,port_bottom,inside_head,inside_bottom)
        prefix = prefixgenerator(port_head, port_bottom,name,sheet)
        code = "/*====================Ports Declaration====================*/\n" + prefix + "\n"
        Variable = Variablegenerator(inside_head, inside_bottom,sheet)
        code = code + \
               "\n/*====================Variable Declaration====================*/\n" + Variable +"\n"
        file =  open("../{}.v".format(name),mode="rt",encoding="utf-8")
        data = file.read()
        pos = re.search("/\*====================Function Code====================\*/", data).span()[0]
        data = code + data[pos:]
        file.close()
        file =  open("../{}.v".format(name),mode="wt",encoding="utf-8")
        print(data)
        file.write(data)
        file.close()

##慎用generatorcode,用之前把代码备份！！！
book = pxl.load_workbook("CPU-26.xlsx")
gl = ["decoder"]
GenerateCode(gl,book)
book.close()