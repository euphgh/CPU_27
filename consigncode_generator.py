from numpy import *

def binary(bit):
    res = []
    for i in range(2**bit):
        temp = list(bin(i))[2:]
        num = len(temp)
        for j in range(bit):
            if j<=num-1:
                temp[j] = int(temp[j])
            else:
                temp.insert(0,0)
        res.append(temp)
    res.sort(key=lambda x:x.count(1))
    res = array(res[1:])
    return res

import openpyxl as pxl
book = pxl.load_workbook("CPU-26.xlsx")
sheet = book.worksheets[0]
instr_h = 31#对应excel 3行
instr_l = 26#对应exce 8行
torow = lambda x:34-x
toitem = lambda x:34-x
sign_l = 35
sign_h = 52
succ = [1 for i in range(sign_l,sign_h+1)]
code = ""
line = sheet.max_column
ena = binary(instr_h-instr_l+1)
for row_sign in range(sign_l,sign_h+1):
    for i in range(len(ena)):
        instruct = {}
        flag = True
        for col_instr in range(2,line+1):
            temp = array([int(sheet.cell(k,col_instr).value) for k in range(torow(instr_h),torow(instr_l)+1)])
            temp = temp*ena[i]
            if instruct.get(temp,2)==2:
                instruct[temp] = sheet.cell(row_sign,col_instr)
            else:
                if insert[temp] != sheet.cell(row_sign,col_instr):
                    flag=False
                    break
        if flag:
            succ.append(1)
            codeline = "assign {} = ".format(sheet.cell(row_sign,1).value)
            for period in instruct.items():#遍历字典中每一个元素
                if instruct[period].value==1:#如果是1的话就生成一个与项
                    anditem = ""
                    for num in range(len(ena[i].count(1))):##查看此时的元素有几个，生成几个元素的与项
                        anditem =anditem + "&&" + "instruct[{}]".format(toitem(num))
                    codeline = codeline+" || "+anditem##生成或项
            code = code + ";\\n"+ codeline
            break
        else:
            succ.append(0)
