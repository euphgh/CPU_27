import openpyxl
import signal
def alwaysgenerator(port_head,port_bottom,sheet):
    forward = ""
    initial = ""
    defines = ""
    for j in range(port_head, port_bottom):

        if sheet.cell(j,4).value == "out" and sheet.cell(j,5).value=="reg":
            forward = forward + "\n\t\t\t"+sheet.cell(j,2).value+" <= {}_wire;".format(sheet.cell(j,2).value[:-4])
            initial = initial + "\n\t\t"+sheet.cell(j,2).value + " <= `ini_{};".format(sheet.cell(j,2).value[:-4])
            defines = defines + "\n" +"`define ini_{0} {1}'b0".format(sheet.cell(j,2).value[:-4],sheet.cell(j,3).value)
    always = '''\
always @(posedge clk ) begin
    if (!rst_n) begin{0}    
    end else begin
        if (forward) begin{1}    
        end
    end
end'''.format(initial,forward)
    print(defines)
    return always

def AlwaysCode(updatelist,book):
    for i in updatelist:
        sheet = book[i]
        name = i
        port_head = 3
        port_bottom = port_head
        for j in range(1, sheet.max_row+1):
            if sheet.cell(j, 1).value == "variable":
                port_bottom = j
                break
        print(alwaysgenerator(port_head,port_bottom,sheet))


book = openpyxl.load_workbook("CPU-26.xlsx")
gl = ["MEM"]
AlwaysCode(gl,book)
book.close()