import numpy as np
import openpyxl as pxl
import re
import copy

book = pxl.load_workbook("CPU-26.xlsx")
sheet = book["sign"]
colmax = pxl.utils.column_index_from_string("AW")
torow = lambda x:34-x

sign_l = 35
sign_h = 89
code = open("signcode.txt","rt",encoding="utf-8")
ansdict = {}
name = ""
flag = True

def testcol(ins,fun,twobit,val):
    if len(twobit) == 0:##递归终止
        return eval(fun)==val
    else:##递归
        temp0 = copy.deepcopy(ins)
        temp0[twobit[0]] = 0
        flag0 = testcol(temp0,fun,twobit[1:],val)
        if flag0 == False:
            return False
        else:
            temp1 = copy.deepcopy(ins)
            temp1[twobit[0]]=1
            flag1 = testcol(temp1,fun,twobit[1:],val)
            if flag1 == False:
                return False
        return True

def test(fun,valist,bit):
    flagtemp = True
    if len(valist)==0:
        print("getvalist error!")
        exit(0)
   
    for col in range(2,colmax+1):
        ins=[]
        temp = copy.deepcopy(bit)
        for i in range(0,32): ##生成列，含有“2”的列表
            tempval = sheet.cell(torow(i),col).value
            if tempval==0 or tempval==1:
                ins.append(tempval)
            else:
                ins.append(2)
        i = 0
        while i<len(temp):
            if ins[temp[i]]!=2:
                temp.remove(temp[i])
                i -= 1
            i += 1
        # print(temp,pxl.utils.get_column_letter(col))
        flagtemp = flagtemp and testcol(ins,fun,temp,valist[col-2])
        # print(col,flagtemp==valist[col-2])        
    return (flagtemp)


def getvalist(name):
    for row in range(sign_l,sign_h+1):
        temp = sheet.cell(row,1).value
        pos = temp.find("/")
        if temp[:pos]==name:
            return np.array([sheet.cell(row,col).value for col in range(2,colmax+1)],dtype=int)
    return np.array([])


def getbit(anditem):
    temp = "\[(\d*)\]"
    ans = re.findall(temp,anditem)
    k = list(map(int,ans))
    return k

funlist = []
line = code.readline()
flag = False if line=="" else True
while flag:
    line = line.replace("&&","&")
    line = line.replace("||"," ")
    line = line.replace("!","1-")
   
    if line[0]=="a":
        andlist = line.strip().split()
        name = andlist[1]
        andlist = andlist[3:]
        if andlist[-1]==";":
            andlist = andlist[:-1]
            funlist.extend(andlist)
            # print(" | ".join(funlist))
            ansdict[name] = test(" | ".join(funlist),getvalist(name),getbit(funlist[0]))
            name = ""
            funlist = []
        else:
            funlist.extend(andlist)
    elif line[0]==" ":
        andlist = line.strip().split()
        if andlist[-1]==";":
            andlist = andlist[:-1]
            funlist.extend(andlist)
            # print(" | ".join(funlist))
            ansdict[name] = test(" | ".join(funlist),getvalist(name),getbit(funlist[0]))
            name = ""
            funlist = []
        else:
            funlist.extend(andlist)
    line = code.readline()
    flag = False if line=="" else True
flag = True

for word in ansdict.items():
    if word[1]==False:
        print("Wrong Expressions name: {}".format(word[0]))
        flag = False
if flag:
    print("All signs expressions are correct!")
else:
    print("Above signs expressions are wrong!")
book.close()